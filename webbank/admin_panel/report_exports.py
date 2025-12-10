import csv
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment

def export_to_csv(filename, headers, data):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    return response

def export_to_pdf(filename, template_path, context):
    template = get_template(template_path)
    html = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
    return HttpResponse('We had some errors <pre>%s</pre>' % html)

def export_to_excel(filename, headers, data, title="Report"):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title

    # Add title
    worksheet.merge_cells('A1:{}1'.format(chr(ord('A') + len(headers) - 1)))
    title_cell = worksheet['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add a blank row for spacing
    worksheet.append([])

    # Add headers
    worksheet.append(headers)
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=worksheet.max_row, column=col_idx)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data
    for row_data in data:
        worksheet.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = worksheet.cell(row=worksheet.max_row, column=col_idx)
            cell.border = thin_border
            # Optional: set column width
            worksheet.column_dimensions[chr(ord('A') + col_idx - 1)].auto_size = True

    workbook.save(response)
    return response
